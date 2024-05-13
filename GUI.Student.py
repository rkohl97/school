import tkinter as tk
from tkinter.ttk import *
from openpyxl import workbook
from openpyxl import load_workbook
import random

# creating tkinter window
root = tk.Tk()
root.title('Lottery')
root.geometry("300x300")
root.resizable(0, 0)
root.grid_columnconfigure(1, weight=1)


# You can use grid_columnconfigure to show empty columns. This shows column 1
# By default, the weight of each column or row is 0, meaning don't expand to fill space.
def randnum():
    wb = load_workbook('hello.xlsx')
    ws = wb.active
    range = ws['A2':'A19']
    l = []
    for cell in range:
        for x in cell:
            l.append(x.value)
    print(l)  # prints all l; not required
    # now I want computer to choose randomly one value out of list l
    computer_action = random.choice(l)
    print(computer_action)  # prints that person's name; not required
    l1.config(text=computer_action)


b1 = tk.Button(text="Select Student Randomly", font=("Arial", 15), bg="#A3E4D7", command=randnum)
b1.grid(row=2, column=1)

l1 = tk.Label(bg="#F39C12", font=("Arial", 10), text="your value will shows up here")
l1.grid(row=4, column=1)

# b1.place(x=110, y=70)
# l1.place(x=165, y=130)

tk.mainloop()
